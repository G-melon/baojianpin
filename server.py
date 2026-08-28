#!/usr/bin/env python3
"""
保健品销售下单系统 — Flask 后端
顾客下单 + 商家后台管理（上传图片/标签/看订单）
"""

import json
import os
import sys
import uuid
import io
import threading
import re
import secrets
import smtplib
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='.', static_url_path='')

# ========== 配置 ==========
DATA_DIR = Path(__file__).parent / 'data'
UPLOAD_DIR = Path(__file__).parent / 'uploads'
PRODUCTS_FILE = DATA_DIR / 'products.json'
ORDERS_FILE = DATA_DIR / 'orders.json'
# SVG 已移除：SVG 内可嵌 <script>，从 /uploads/ 同源返回会触发存储型 XSS
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_UPLOAD_SIZE = 5 * 1024 * 1024  # 5 MB
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_SIZE
ORDER_STATUSES = ["新订单", "已联系", "已确认", "生产中", "已发货", "已完成", "已取消"]
DEFAULT_ORDER_STATUS = "新订单"
ORDER_EDIT_WINDOW_SECONDS = 10 * 60

# JSON 读改写需要互斥，否则两个顾客同时下单会丢单
_data_lock = threading.Lock()
_admin_sessions = {}
ADMIN_CREDENTIALS_FILE = DATA_DIR / 'admin_credentials.json'

DATA_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)

def env_value(name, default=''):
    return os.environ.get(name, default).strip()

# ========== 初始产品数据 ==========
# 保留产品管理框架，默认不预置任何商品。
DEFAULT_PRODUCTS = []

CATEGORIES = {
    "全部": {"name": "全部"},
    "营养补充": {"name": "营养补充"},
    "维生素矿物质": {"name": "维生素矿物质"},
    "膳食纤维": {"name": "膳食纤维"},
    "益生菌调理": {"name": "益生菌调理"},
    "心脑血管": {"name": "心脑血管"},
    "儿童成长": {"name": "儿童成长"},
    "抗炎护肝": {"name": "抗炎护肝"},
    "中老年营养": {"name": "中老年营养"},
    "美容养护": {"name": "美容养护"},
    "运动营养": {"name": "运动营养"},
    "礼盒套装": {"name": "礼盒套装"}
}
CATEGORY_NAMES = [k for k in CATEGORIES.keys() if k != "全部"]
CATEGORY_ALIASES = {
    "vitamin": "维生素矿物质",
    "fiber": "膳食纤维",
    "probiotic": "益生菌调理",
    "cardio": "心脑血管",
    "kids": "儿童成长",
    "liver": "抗炎护肝",
    "senior": "中老年营养",
    "beauty": "美容养护",
    "sports": "运动营养",
    "gift": "礼盒套装"
}
CATEGORY_KEYWORDS = [
    ("抗炎护肝", ("姜黄", "护肝", "抗炎", "美藤果油")),
    ("益生菌调理", ("益生菌", "菌株", "肠道", "口腔", "幽门", "睡眠益生菌", "女性益生菌", "呼吸道")),
    ("心脑血管", ("鱼油", "辅酶", "Q10", "降血脂", "心脏")),
    ("美容养护", ("胶原", "透明质酸", "美白", "抗氧", "护肤", "养颜", "虾青素")),
    ("膳食纤维", ("膳食纤维", "菊粉", "益生元", "纤维", "代餐", "控糖", "饱腹")),
    ("维生素矿物质", ("维生素", "矿物质", "钙", "铁", "锌", "硒", "叶酸", "复合维生素")),
    ("儿童成长", ("DHA", "学童", "儿童", "记忆力", "专注")),
    ("中老年营养", ("中老年", "老人", "骨骼", "关节", "钙片", "辅酶")),
    ("运动营养", ("运动", "蛋白", "乳清", "肌酸", "健身")),
    ("礼盒套装", ("礼盒", "套装", "组合", "年礼", "送礼")),
    ("营养补充", ("营养", "保健", "健康", "补充"))
]

BG_OPTIONS = [
    {"value":"red-bg",    "label":"经典红"},
    {"value":"gold-bg",   "label":"富贵金"},
    {"value":"pink-bg",   "label":"浪漫粉"},
    {"value":"purple-bg", "label":"典雅紫"},
    {"value":"blue-bg",   "label":"深海蓝"},
    {"value":"green-bg",  "label":"翠竹绿"},
    {"value":"orange-bg", "label":"活力橙"},
    {"value":"teal-bg",   "label":"青玉色"},
    {"value":"rose-bg",   "label":"玫瑰红"},
    {"value":"indigo-bg", "label":"靛青蓝"}
]

# ========== 数据读写 ==========
# load/save 不加锁，纯单读单写的工具函数；任何「读→改→写」序列必须用 mutate_json()。
def load_json(path, default=None):
    if path.exists():
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return default if default is not None else []

def save_json(path, data):
    # 写到临时文件再原子 rename，避免半写文件
    tmp = path.with_suffix(path.suffix + '.tmp')
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def mutate_json(path, mutator, default=None):
    """加锁的 load→mutate→save。mutator(data) 返回新 data 与可选的额外结果。"""
    with _data_lock:
        data = load_json(path, default if default is not None else [])
        result = mutator(data)
        if isinstance(result, tuple):
            new_data, extra = result
        else:
            new_data, extra = result, None
        save_json(path, new_data)
        return extra

def infer_category(name='', desc='', tags=None):
    """根据后台输入的中文内容自动判断产品主分类。"""
    tags = tags or []
    text = ' '.join([str(name or ''), str(desc or ''), ' '.join(str(t) for t in tags)])
    for category, keywords in CATEGORY_KEYWORDS:
        if any(keyword in text for keyword in keywords):
            return category
    return "营养补充"

def normalize_category(cat, name='', desc='', tags=None):
    if cat in CATEGORY_NAMES:
        return cat
    if cat in CATEGORY_ALIASES:
        return CATEGORY_ALIASES[cat]
    return infer_category(name, desc, tags)

def normalize_product(product):
    tags = product.get('tags') or []
    if not isinstance(tags, list):
        tags = [str(tags)]
    clean_tags = []
    for tag in tags:
        tag = str(tag).strip()
        if tag and tag not in clean_tags:
            clean_tags.append(tag)
    product['tags'] = clean_tags
    product['cat'] = normalize_category(
        product.get('cat', ''),
        product.get('name', ''),
        product.get('desc', ''),
        product['tags']
    )
    raw_card_options = product.get('cardOptions')
    if not isinstance(raw_card_options, list):
        raw_card_options = re.split(r'[、,，;；\n]+', str(product.get('cardCount') or ''))
    card_options = []
    for option in raw_card_options:
        option = str(option or '').strip()
        if option and option not in card_options:
            card_options.append(option)
    product['cardOptions'] = card_options
    product['cardCount'] = '、'.join(card_options)
    raw_product_options = product.get('productOptions')
    if not isinstance(raw_product_options, list):
        raw_product_options = re.split(r'[、,，;；\n]+', str(product.get('productOptionText') or ''))
    product_options = []
    for option in raw_product_options:
        option = str(option or '').strip()
        if option and option not in product_options:
            product_options.append(option)
    product['productOptions'] = product_options
    product['productOptionText'] = '、'.join(product_options)
    return product

def normalize_order(order):
    if order.get("status") not in ORDER_STATUSES:
        order["status"] = DEFAULT_ORDER_STATUS
    return order

def public_order(order):
    result = normalize_order(dict(order))
    result.pop('editToken', None)
    return result

def order_edit_deadline(order):
    raw = order.get('timestamp')
    try:
        created_at = datetime.fromisoformat(str(raw))
    except (TypeError, ValueError):
        return None
    return created_at.timestamp() + ORDER_EDIT_WINDOW_SECONDS

def can_customer_edit(order):
    deadline = order_edit_deadline(order)
    return deadline is not None and datetime.now().timestamp() <= deadline

def sanitize_order_payload(data):
    if not data or not data.get('customerName'):
        return "请填写下单人姓名", None
    if not data.get('customerPhone'):
        return "请填写联系电话", None
    if not data.get('customerAddress'):
        return "请填写收货地址", None
    if not data.get('items') or len(data['items']) == 0:
        return "请选择商品", None
    items = []
    for item in data.get('items', []):
        try:
            qty = int(float(item.get('qty', 0) or 0))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        items.append({
            "id": str(item.get('id', '')).strip(),
            "name": str(item.get('name', '')).strip(),
            "emoji": str(item.get('emoji', '')).strip(),
            "cat": str(item.get('cat', '')).strip(),
            "cardCount": str(item.get('cardCount', '')).strip(),
            "productOption": str(item.get('productOption', '')).strip(),
            "internalPrice": item.get('internalPrice', ''),
            "price": item.get('price', ''),
            "qty": qty
        })
    if not items:
        return "请选择商品", None
    return None, {
        "customerName": data['customerName'].strip(),
        "customerPhone": data.get('customerPhone', '').strip(),
        "customerEmail": data.get('customerEmail', '').strip(),
        "customerAddress": data.get('customerAddress', '').strip(),
        "note": data.get('note', '').strip(),
        "items": items,
        "totalQty": sum(it.get('qty', 0) for it in items)
    }

def format_order_notification(order):
    lines = [
        '保健品家庭内购平台有新订单',
        '',
        f'订单编号：{order.get("orderNo", "")}',
        f'下单人：{order.get("customerName", "")}',
        f'联系电话：{order.get("customerPhone", "")}',
        f'收货地址：{order.get("customerAddress", "")}',
        f'下单日期：{order.get("date", "")}',
    ]
    if order.get('customerEmail'):
        lines.append(f'邮箱：{order.get("customerEmail")}')
    if order.get('note'):
        lines.extend(['', f'客户备注：{order.get("note")}'])
    lines.extend(['', '商品明细：'])
    total_qty = 0
    total_amount = 0
    for item in order.get('items', []):
        qty = item.get('qty', 0) or 0
        total_qty += qty
        try:
            price = float(item.get('internalPrice') or item.get('price') or 0)
        except (TypeError, ValueError):
            price = 0
        total_amount += price * qty
        spec = item.get('cardCount') or item.get('cardOption') or item.get('spec') or ''
        option = item.get('productOption') or ''
        detail = ' / '.join(str(v) for v in (spec, option) if v)
        price_text = f'，单价¥{price:g}' if price else ''
        lines.append(f'- {item.get("name", "")} x {qty}{price_text}' + (f'（{detail}）' if detail else ''))
    lines.extend(['', f'总件数：{total_qty} 件'])
    if total_amount:
        lines.append(f'预估总金额：¥{total_amount:g}')
    return '\n'.join(lines)

def send_order_notification(order):
    smtp_host = env_value('SMTP_HOST')
    smtp_user = env_value('SMTP_USER')
    smtp_password = os.environ.get('SMTP_PASSWORD', '')
    notify_to = env_value('ORDER_NOTIFY_TO') or env_value('SMTP_TO')
    if not (smtp_host and smtp_user and smtp_password and notify_to):
        return

    smtp_port = int(env_value('SMTP_PORT', '465') or 465)
    smtp_from = env_value('SMTP_FROM') or smtp_user
    subject = f'新订单：{order.get("customerName", "")} {order.get("orderNo", "")}'
    message = EmailMessage()
    message['Subject'] = subject
    message['From'] = smtp_from
    message['To'] = notify_to
    message.set_content(format_order_notification(order))

    def _send():
        try:
            if smtp_port == 465:
                with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=10) as server:
                    server.login(smtp_user, smtp_password)
                    server.send_message(message)
            else:
                with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
                    server.starttls()
                    server.login(smtp_user, smtp_password)
                    server.send_message(message)
        except Exception as exc:
            print(f'Order notification email failed: {exc}', file=sys.stderr)

    threading.Thread(target=_send, daemon=True).start()

def create_admin_session(account):
    token = secrets.token_urlsafe(32)
    _admin_sessions[token] = {
        "account": account,
        "createdAt": datetime.now().isoformat()
    }
    return token

def get_admin_credentials():
    account = os.environ.get('ADMIN_ACCOUNT', '').strip()
    password = os.environ.get('ADMIN_PASSWORD', '')
    if account and password:
        return account, password
    if ADMIN_CREDENTIALS_FILE.exists():
        try:
            with open(ADMIN_CREDENTIALS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return str(data.get('account', '')).strip(), str(data.get('password', ''))
        except (OSError, json.JSONDecodeError):
            return '', ''
    return '', ''

# ========== 静态文件 ==========
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/admin')
def admin():
    return send_from_directory('.', 'admin.html')

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(str(UPLOAD_DIR), filename)

# ========== 产品 API ==========
@app.route('/api/products', methods=['GET'])
def get_products():
    products = load_json(PRODUCTS_FILE, DEFAULT_PRODUCTS)
    products = [normalize_product(dict(p)) for p in products]
    return jsonify(products)

@app.route('/api/product-sales', methods=['GET'])
def get_product_sales():
    """公开销量累计：按已提交订单里的商品 id/name 统计购买件数。"""
    orders = load_json(ORDERS_FILE, [])
    sales = {}
    for order in orders:
        for item in order.get('items', []):
            try:
                qty = int(float(item.get('qty', 0) or 0))
            except (TypeError, ValueError):
                qty = 0
            for key in (item.get('id'), item.get('name')):
                key = str(key or '').strip()
                if key:
                    sales[key] = sales.get(key, 0) + qty
    return jsonify(sales)

@app.route('/api/products', methods=['POST'])
def add_product():
    """添加新产品（需要 admin 鉴权）"""
    if not check_admin():
        return jsonify({"error":"未授权"}), 401
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({"error":"产品名称不能为空"}), 400
    tags = data.get('tags', [])
    if not isinstance(tags, list):
        tags = []
    product = normalize_product({
        "id": 'p' + uuid.uuid4().hex[:8],
        "cat": data.get('cat', ''),
        "name": data['name'],
        "desc": data.get('desc', ''),
        "emoji": data.get('emoji', '🎁'),
        "bg": data.get('bg', 'red-bg'),
        "tags": tags,
        "efficacy": data.get('efficacy', ''),
        "usage": data.get('usage', ''),
        "internalPrice": data.get('internalPrice', ''),
        "internalPriceText": data.get('internalPriceText', ''),
        "details": data.get('details', []),
        "usageDetails": data.get('usageDetails', []),
        "specs": data.get('specs', {}),
        "cardCount": data.get('cardCount', ''),
        "cardOptions": data.get('cardOptions', []),
        "productOptionText": data.get('productOptionText', ''),
        "productOptions": data.get('productOptions', []),
        "image": data.get('image', None),
        "images": data.get('images', [])
    })
    def _add(products):
        products.append(product)
        return products
    mutate_json(PRODUCTS_FILE, _add, DEFAULT_PRODUCTS)
    return jsonify(product), 201

@app.route('/api/products/<product_id>', methods=['PUT'])
def update_product(product_id):
    """更新产品"""
    if not check_admin():
        return jsonify({"error":"未授权"}), 401
    data = request.get_json()
    def _update(products):
        for p in products:
            if p['id'] == product_id:
                for k in ('name','cat','desc','emoji','bg','tags','image','images','efficacy','usage','internalPrice','internalPriceText','details','usageDetails','specs','cardCount','cardOptions','productOptionText','productOptions'):
                    if k in data: p[k] = data[k]
                normalize_product(p)
                return products, p
        return products, None
    updated = mutate_json(PRODUCTS_FILE, _update, DEFAULT_PRODUCTS)
    if updated is None:
        return jsonify({"error":"产品不存在"}), 404
    return jsonify(updated)

@app.route('/api/products/<product_id>', methods=['DELETE'])
def delete_product(product_id):
    """删除产品"""
    if not check_admin():
        return jsonify({"error":"未授权"}), 401
    mutate_json(PRODUCTS_FILE,
                lambda products: [p for p in products if p['id'] != product_id],
                DEFAULT_PRODUCTS)
    return jsonify({"ok":True})

# ========== 图片上传 ==========
@app.route('/api/upload', methods=['POST'])
def upload_image():
    """上传产品图片"""
    if not check_admin():
        return jsonify({"error":"未授权"}), 401
    if 'file' not in request.files:
        return jsonify({"error":"未选择文件"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error":"文件名为空"}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error":f"不支持的文件类型: .{ext}（仅支持 png/jpg/jpeg/gif/webp）"}), 400
    # 验证文件实际内容是图片（防止改后缀绕过）
    file.stream.seek(0)
    head = file.stream.read(16)
    file.stream.seek(0)
    image_signatures = [
        b'\x89PNG\r\n\x1a\n',     # png
        b'\xff\xd8\xff',          # jpg/jpeg
        b'GIF87a', b'GIF89a',     # gif
        b'RIFF',                  # webp（前 4 字节，后续 8 字节是 size + 'WEBP'）
    ]
    if not any(head.startswith(sig) for sig in image_signatures):
        return jsonify({"error":"文件内容不是有效的图片"}), 400
    filename = 'img_' + uuid.uuid4().hex[:12] + '.' + ext
    file.save(str(UPLOAD_DIR / filename))
    return jsonify({"url": f"/uploads/{filename}", "filename": filename})

# ========== 订单 API ==========
@app.route('/api/orders', methods=['GET'])
def get_orders():
    """获取订单列表（admin 鉴权）"""
    if not check_admin():
        return jsonify({"error":"未授权"}), 401
    orders = load_json(ORDERS_FILE, [])
    orders = [public_order(o) for o in orders]
    return jsonify(orders)

@app.route('/api/orders', methods=['POST'])
def submit_order():
    """提交新订单（顾客端无需鉴权）"""
    data = request.get_json()
    error, clean = sanitize_order_payload(data)
    if error:
        return jsonify({"error": error}), 400
    now = datetime.now()
    order = {
        "orderNo": 'HP' + now.strftime('%Y%m%d%H%M%S') + uuid.uuid4().hex[:4],
        "customerName": clean["customerName"],
        "customerPhone": clean["customerPhone"],
        "customerEmail": clean["customerEmail"],
        "customerAddress": clean["customerAddress"],
        "note": clean["note"],
        "date": now.strftime('%Y/%m/%d'),
        "timestamp": now.isoformat(),
        "status": DEFAULT_ORDER_STATUS,
        "editToken": secrets.token_urlsafe(24),
        "items": clean["items"],
        "totalQty": clean["totalQty"]
    }
    def _insert(orders):
        orders.insert(0, order)
        return orders
    mutate_json(ORDERS_FILE, _insert, [])
    send_order_notification(order)
    return jsonify(order), 201

@app.route('/api/orders/<order_no>', methods=['GET'])
def get_customer_order(order_no):
    """顾客凭编辑 token 查看自己的订单。"""
    token = request.args.get('token', '').strip()
    orders = load_json(ORDERS_FILE, [])
    for order in orders:
        if order.get("orderNo") == order_no and order.get("editToken") == token:
            result = public_order(order)
            result["editable"] = can_customer_edit(order)
            result["editDeadline"] = order_edit_deadline(order)
            return jsonify(result)
    return jsonify({"error":"订单不存在或凭证无效"}), 404

@app.route('/api/orders/<order_no>', methods=['PUT'])
def update_customer_order(order_no):
    """顾客 10 分钟内凭编辑 token 修改自己的订单。"""
    data = request.get_json() or {}
    token = str(data.get('editToken', '')).strip()
    error, clean = sanitize_order_payload(data)
    if error:
        return jsonify({"error": error}), 400

    def _update(orders):
        for order in orders:
            if order.get("orderNo") != order_no:
                continue
            if order.get("editToken") != token:
                return orders, ("invalid", None)
            if not can_customer_edit(order):
                return orders, ("expired", None)
            order.update(clean)
            order["updatedAt"] = datetime.now().isoformat()
            return orders, ("ok", order)
        return orders, ("missing", None)

    status, updated = mutate_json(ORDERS_FILE, _update, [])
    if status == "invalid":
        return jsonify({"error":"订单凭证无效"}), 403
    if status == "expired":
        return jsonify({"error":"订单已超过10分钟，不能修改"}), 409
    if status == "missing":
        return jsonify({"error":"订单不存在"}), 404
    send_order_notification(updated)
    return jsonify(updated)

@app.route('/api/orders/<order_no>/status', methods=['PUT'])
def update_order_status(order_no):
    """更新订单状态（admin 鉴权）"""
    if not check_admin():
        return jsonify({"error":"未授权"}), 401
    data = request.get_json() or {}
    status = data.get("status", "")
    if status not in ORDER_STATUSES:
        return jsonify({"error":"订单状态无效"}), 400

    def _update(orders):
        for order in orders:
            if order.get("orderNo") == order_no:
                order["status"] = status
                order["statusUpdatedAt"] = datetime.now().isoformat()
                return orders, normalize_order(order)
        return orders, None

    updated = mutate_json(ORDERS_FILE, _update, [])
    if updated is None:
        return jsonify({"error":"订单不存在"}), 404
    return jsonify(updated)

@app.route('/api/orders', methods=['DELETE'])
def clear_orders():
    """清空订单"""
    if not check_admin():
        return jsonify({"error":"未授权"}), 401
    with _data_lock:
        save_json(ORDERS_FILE, [])
    return jsonify({"ok":True})

# ========== XLSX 导出 ==========
def filtered_orders_from_request():
    orders = load_json(ORDERS_FILE, [])
    name_filter = request.args.get('name', '').strip().lower()
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    order_no = request.args.get('orderNo', '').strip()

    filtered = []
    for o in orders:
        if order_no and o['orderNo'] != order_no:
            continue
        if name_filter and name_filter not in o['customerName'].lower():
            continue
        if date_from and o['date'] < date_from.replace('-', '/'):
            continue
        if date_to and o['date'] > date_to.replace('-', '/'):
            continue
        filtered.append(o)
    return filtered, name_filter, order_no

def build_orders_workbook(filtered, include_prices=False):
    # 按客户分组（保持订单原顺序）
    customer_orders = {}
    for o in filtered:
        name = o['customerName']
        customer_orders.setdefault(name, []).append(o)

    products = load_json(PRODUCTS_FILE, [])
    product_lookup = {}
    for product in products:
        product_lookup[str(product.get('id', ''))] = product
        product_lookup[str(product.get('name', ''))] = product

    def item_unit_price(item):
        raw = item.get('internalPrice', item.get('price', ''))
        if raw in ('', None):
            product = product_lookup.get(str(item.get('id', ''))) or product_lookup.get(str(item.get('name', '')))
            raw = product.get('internalPrice', '') if product else ''
        try:
            return float(raw)
        except (TypeError, ValueError):
            return None

    def numeric_qty(item):
        try:
            return float(item.get('qty', 0))
        except (TypeError, ValueError):
            return 0

    def format_amount(value):
        return str(int(value)) if float(value).is_integer() else f'{value:.2f}'

    def order_address(order):
        for key in ('customerAddress', 'address', 'shippingAddress', 'receiverAddress', '收货地址'):
            value = order.get(key)
            if value:
                return str(value).strip()
        return ''

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name='宋体', size=20, bold=True)
    info_font  = Font(name='宋体', size=13, bold=True)
    head_font  = Font(name='宋体', size=12, bold=True)
    body_font  = Font(name='宋体', size=11, bold=True)
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def apply_border_block(ws, r1, r2, c1, c2):
        """在合并 *之前* 给每个单元格刷边框，否则 MergedCell 拿不到边框、打印线缺失。"""
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(row=r, column=c).border = border

    for cust_name, cust_orders in customer_orders.items():
        # Excel sheet 名 31 字符上限 + 不允许的字符替换
        safe_name = ''.join(ch if ch not in '[]:*?/\\' else '_' for ch in cust_name)[:31] or '客户'
        ws = wb.create_sheet(title=safe_name)

        # 列宽：A:B 商品名称 / C 规格 / D 数量 / E:F 单位 / G:H 单价 / I 金额
        for col, w in [('A',16),('B',18),('C',12),('D',10),('E',6),('F',6),('G',8),('H',8),('I',14)]:
            ws.column_dimensions[col].width = w

        # ---- Row 1-2: 客户名称 / 发货日期 ----
        apply_border_block(ws, 1, 2, 1, 9)
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:F2')
        ws.merge_cells('G1:I2')
        ws['A1'] = '客户名称';        ws['A1'].font = title_font; ws['A1'].alignment = center_wrap
        ws['B1'] = cust_name;         ws['B1'].font = title_font; ws['B1'].alignment = center_wrap
        ws['G1'] = f'发货日期：{cust_orders[0].get("date","")}'
        ws['G1'].font = info_font;    ws['G1'].alignment = center_wrap
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 26

        # ---- Row 3-4: 总件数 + 订单编号 ----
        total_qty = sum(sum(it.get('qty',0) for it in o.get('items',[])) for o in cust_orders)
        order_nos = '、'.join(o.get('orderNo','') for o in cust_orders)
        apply_border_block(ws, 3, 4, 1, 9)
        ws.merge_cells('A3:I4')
        phones = '、'.join(sorted({o.get('customerPhone', '') for o in cust_orders if o.get('customerPhone')}))
        emails = '、'.join(sorted({o.get('customerEmail', '') for o in cust_orders if o.get('customerEmail')}))
        addresses = '、'.join(sorted({order_address(o) for o in cust_orders if order_address(o)}))
        customer_total = sum(
            (item_unit_price(item) or 0) * numeric_qty(item)
            for order in cust_orders
            for item in order.get('items', [])
        )
        contact_text = f'  总件数：{total_qty} 件     订单编号：{order_nos}'
        if phones:
            contact_text += f'     电话：{phones}'
        if emails:
            contact_text += f'     邮箱：{emails}'
        if include_prices:
            contact_text += f'     总金额：¥{format_amount(customer_total)}'
        ws['A3'] = contact_text
        ws['A3'].font = info_font; ws['A3'].alignment = left_wrap
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 22

        # ---- Row 5-6: 物流信息 + 收货地址 ----
        apply_border_block(ws, 5, 6, 1, 9)
        ws.merge_cells('A5:C6')
        ws.merge_cells('D5:I6')
        ws['A5'] = '快递/物流：\n单号：'
        ws['A5'].font = info_font; ws['A5'].alignment = left_wrap
        ws['D5'] = f'收货地址：{addresses or "未填写"}'
        ws['D5'].font = info_font; ws['D5'].alignment = left_wrap
        ws.row_dimensions[5].height = 32
        ws.row_dimensions[6].height = 32

        # ---- Row 7-8: 列标题 ----
        apply_border_block(ws, 7, 8, 1, 9)
        ws.merge_cells('A7:B8')
        ws.merge_cells('C7:C8')
        ws.merge_cells('D7:D8')
        ws.merge_cells('E7:F8')
        ws.merge_cells('G7:H8')
        ws.merge_cells('I7:I8')
        headings = [(1,'商品名称'),(3,'规格'),(4,'数量'),(5,'单位')]
        if include_prices:
            headings += [(7,'单价'),(9,'金额')]
        for col, txt in headings:
            c = ws.cell(row=7, column=col, value=txt)
            c.font = head_font; c.alignment = center_wrap
        ws.row_dimensions[7].height = 22
        ws.row_dimensions[8].height = 22

        # ---- 数据行：每个商品占 2 行合并块；模板默认 14 个槽位，不够就往下延 ----
        all_items = [it for o in cust_orders for it in o.get('items', [])]
        slot_count = max(14, len(all_items))

        r = 9
        for i in range(slot_count):
            apply_border_block(ws, r, r+1, 1, 9)
            ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=2)  # A:B 商品名称
            ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)  # C 规格
            ws.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=4)  # D 数量
            ws.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=6)  # E:F 单位
            ws.merge_cells(start_row=r, start_column=7, end_row=r+1, end_column=8)  # G:H 单价（留空）
            ws.merge_cells(start_row=r, start_column=9, end_row=r+1, end_column=9)  # I 金额（留空）

            if i < len(all_items):
                item = all_items[i]
                unit_price = item_unit_price(item) if include_prices else None
                qty = item.get('qty','')
                ws.cell(row=r, column=1, value=item.get('name','')).font = body_font
                ws.cell(row=r, column=1).alignment = center_wrap
                ws.cell(row=r, column=3, value=item.get('cardCount','')).font = body_font
                ws.cell(row=r, column=3).alignment = center_wrap
                ws.cell(row=r, column=4, value=qty).font = body_font
                ws.cell(row=r, column=4).alignment = center_wrap
                ws.cell(row=r, column=5, value='件').font = body_font
                ws.cell(row=r, column=5).alignment = center_wrap
                if unit_price is not None:
                    amount = unit_price * numeric_qty(item)
                    ws.cell(row=r, column=7, value=unit_price).font = body_font
                    ws.cell(row=r, column=7).alignment = center_wrap
                    ws.cell(row=r, column=9, value=amount).font = body_font
                    ws.cell(row=r, column=9).alignment = center_wrap
            ws.row_dimensions[r].height = 20
            ws.row_dimensions[r+1].height = 20
            r += 2

        # ---- Footer 行：对应模板 row 37 / row 38（A:B / E:F / G:H 各占 1 行合并） ----
        for _ in range(2):
            apply_border_block(ws, r, r, 1, 9)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
            ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
            ws.row_dimensions[r].height = 22
            r += 1

        # 打印设置：A4 纵向，水平居中
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = 9
        ws.print_options.horizontalCentered = True
        ws.print_area = f'A1:I{r-1}'

    return wb, customer_orders

def workbook_bytes(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# 模板：data 行为 2 行合并块，列 A:B 商品名称 / C 规格 / D 数量 / E:F 单位 / G:H 单价 / I 金额。
@app.route('/api/orders/export-xlsx', methods=['GET'])
def export_orders_xlsx():
    """按客户分 sheet 导出单份 XLSX。mode=bill 为账单，mode=list 为不含价格商品清单。"""
    if not check_admin():
        return jsonify({"error":"未授权"}), 401

    filtered, name_filter, order_no = filtered_orders_from_request()
    if not filtered:
        return jsonify({"error":"当前筛选无订单"}), 404

    mode = request.args.get('mode', 'list').strip().lower()
    include_prices = mode == 'bill'
    wb, customer_orders = build_orders_workbook(filtered, include_prices=include_prices)
    output = workbook_bytes(wb)

    if order_no:
        cust = next(iter(customer_orders.keys()))
        suffix = f'{cust}_{order_no}'
    else:
        suffix = name_filter or '全部'
    label = '账单' if include_prices else '商品清单'
    filename = f'保健品{label}_{suffix}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

# ========== Admin 鉴权 ==========
@app.route('/api/admin/state', methods=['GET'])
def admin_state():
    """返回后台登录方式。"""
    return jsonify({"authMode": "fixed_password"})

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    data = request.get_json() or {}
    account = str(data.get('account', '')).strip()
    password = str(data.get('password', ''))
    expected_account, expected_password = get_admin_credentials()
    if not expected_account or not expected_password:
        return jsonify({"error": "后台账号未配置"}), 503
    if not (
        secrets.compare_digest(account, expected_account) and
        secrets.compare_digest(password, expected_password)
    ):
        return jsonify({"error": "账号或密码错误"}), 403
    return jsonify({"token": create_admin_session(account), "ok": True, "account": account})

@app.route('/api/admin/check', methods=['GET'])
def admin_check():
    """检查是否已登录"""
    token = request.headers.get('X-Admin-Token', '')
    if token in _admin_sessions:
        return jsonify({"ok":True})
    return jsonify({"ok":False}), 401

@app.route('/api/categories', methods=['GET'])
def get_categories():
    """获取分类列表"""
    return jsonify(CATEGORIES)

@app.route('/api/bg-options', methods=['GET'])
def get_bg_options():
    """获取背景色选项"""
    return jsonify(BG_OPTIONS)

def check_admin():
    """检查 admin 鉴权"""
    token = request.headers.get('X-Admin-Token', '')
    return token in _admin_sessions

# ========== 启动 ==========
if __name__ == '__main__':
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding='utf-8')
    # debug 默认关闭：Werkzeug 调试器暴露在 0.0.0.0 会让任何人远程执行 Python
    # 本地开发时用：FLASK_DEBUG=1 python3 server.py
    debug = os.environ.get('FLASK_DEBUG', '').lower() in ('1', 'true', 'yes')
    print("保健品家庭内购平台后端启动中...")
    print("   顾客端: http://localhost:8899")
    print("   商家后台: http://localhost:8899/admin")
    if debug:
        print("   ⚠️ DEBUG 已开启，仅供本机调试，不要在公网暴露！")
    app.run(host='0.0.0.0', port=8899, debug=debug)




