import json
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import openpyxl


SOURCE = Path(
    r"C:\Users\cjbbe\xwechat_files\wxid_yvsowc6f9b3122_e1f2\msg\file\2026-08\内购产品明细--20260820.xlsx"
)
ROOT = Path(__file__).resolve().parents[1]
PRODUCTS_FILE = ROOT / "data" / "products.json"
IMAGE_DIR = ROOT / "uploads" / "products"


def slugify(value):
    value = re.sub(r"[®™]", "", str(value))
    value = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value).strip("-").lower()
    return value or "product"


def text_lines(value):
    return [line.strip() for line in str(value or "").splitlines() if line.strip()]


def image_id(formula):
    match = re.search(r'"(ID_[A-F0-9]+)"', str(formula or ""))
    return match.group(1) if match else ""


def parse_cell_images(workbook_path):
    with zipfile.ZipFile(workbook_path) as archive:
        rel_root = ET.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib.get("Target")
            for rel in rel_root
        }
        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
        }
        img_root = ET.fromstring(archive.read("xl/cellimages.xml"))
        id_to_media = {}
        for cell_image in img_root.findall("etc:cellImage", ns):
            c_nv_pr = cell_image.find(".//xdr:cNvPr", ns)
            blip = cell_image.find(".//a:blip", ns)
            if c_nv_pr is None or blip is None:
                continue
            rel_id = blip.attrib.get(f"{{{ns['r']}}}embed")
            target = rid_to_target.get(rel_id)
            if target and target != "NULL":
                id_to_media[c_nv_pr.attrib.get("name")] = "xl/" + target
        return id_to_media


def infer_category(name, efficacy):
    text = f"{name} {efficacy}"
    rules = [
        ("抗炎护肝", ["姜黄", "护肝", "抗炎", "美藤果油"]),
        ("益生菌调理", ["益生菌", "菌株", "肠道", "口腔", "幽门", "睡眠益生菌", "女性益生菌", "呼吸道"]),
        ("心脑血管", ["鱼油", "辅酶", "Q10", "降血脂", "心脏"]),
        ("美容养护", ["胶原", "透明质酸", "虾青素", "美容", "皮肤", "淡斑"]),
        ("膳食纤维", ["菊粉", "益生元", "纤维"]),
        ("维生素矿物质", ["维生素", "钙", "维生素D", "维生素K", "AD"]),
        ("儿童成长", ["DHA", "学童", "儿童", "记忆力", "专注"]),
    ]
    for category, words in rules:
        if any(word in text for word in words):
            return category
    return "营养补充"


def bg_for(category):
    return {
        "益生菌调理": "green-bg",
        "心脑血管": "red-bg",
        "维生素矿物质": "gold-bg",
        "儿童成长": "blue-bg",
        "抗炎护肝": "orange-bg",
        "美容养护": "pink-bg",
        "膳食纤维": "teal-bg",
    }.get(category, "gold-bg")


def date_text(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value or "")


def main():
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    id_to_media = parse_cell_images(SOURCE)
    products = []
    used_slugs = {}

    workbook = openpyxl.load_workbook(SOURCE, data_only=False)
    sheet = workbook["260820-实"]

    with zipfile.ZipFile(SOURCE) as archive:
        for row_num in range(3, sheet.max_row + 1):
            name = sheet.cell(row_num, 2).value
            if not name:
                continue
            efficacy = sheet.cell(row_num, 3).value or ""
            usage = sheet.cell(row_num, 4).value or ""
            internal_price = sheet.cell(row_num, 5).value
            production_date = sheet.cell(row_num, 7).value
            stock = sheet.cell(row_num, 8).value

            base_slug = slugify(name)
            used_slugs[base_slug] = used_slugs.get(base_slug, 0) + 1
            slug = base_slug if used_slugs[base_slug] == 1 else f"{base_slug}-{used_slugs[base_slug]}"

            media = id_to_media.get(image_id(sheet.cell(row_num, 6).value))
            image_url = None
            if media and media in archive.namelist():
                ext = Path(media).suffix.lower() or ".png"
                image_path = IMAGE_DIR / f"{slug}{ext}"
                image_path.write_bytes(archive.read(media))
                image_url = f"/uploads/products/{image_path.name}"

            category = infer_category(name, efficacy)
            usage_lines = text_lines(usage)
            spec = usage_lines[0] if usage_lines else ""
            price_text = "" if internal_price in (None, "") else f"¥{internal_price}/瓶"
            stock_text = "" if stock in (None, "") else str(stock)
            product = {
                "id": slug,
                "cat": category,
                "name": name,
                "desc": (text_lines(efficacy)[0] if text_lines(efficacy) else "")[:80],
                "emoji": "",
                "bg": bg_for(category),
                "tags": [category],
                "image": image_url,
                "images": [image_url] if image_url else [],
                "efficacy": efficacy,
                "usage": usage,
                "internalPrice": internal_price,
                "internalPriceText": price_text,
                "productionDate": date_text(production_date),
                "stock": stock_text,
                "code": f"NG-{row_num - 2:03d}",
                "cardCount": spec,
                "cardOptions": [spec] if spec else [],
                "productOptionText": "",
                "productOptions": [],
                "specs": {
                    "商品名称": name,
                    "规格": spec,
                    "内购价": price_text,
                    "生产日期": date_text(production_date),
                    "库存": stock_text,
                },
                "details": text_lines(efficacy),
                "usageDetails": usage_lines,
            }
            products.append(product)

    PRODUCTS_FILE.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "products": len(products),
        "images": sum(1 for product in products if product.get("image")),
        "categories": sorted({product["cat"] for product in products}),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
